# Перед началом работы нужно положить рядом файл "Оплата агентам за месяц.xlx"

from __future__ import print_fuction
from mailmerge import MailMerge
from datetime import date
from openpyxl import load_workbook
from calendar import monthrange
import pandas as pd
import pyodbc
import datetime

# Функция уменьшиения отображения количества знаков после запятой
def toFix(Obj,n=0):
    return f"{Obj:.{n}f}"

# Функция проверки корректности даты
def datetry(date_test):
    try:
        if date_test != datetime.datetime.strptime(date_test,"%d.%m.%Y").strftime("%d.%m.%Y"):
            raise ValueError
        return True
    except ValueError:
        return False


# Функция поиска оператора
def search(oper):
    for row in user:
        line = row
        pfio, name_work, dog_n, date_beg, date_end = [line[0],line[1],line[2],\
                                                      line[3].strftime("%d.%m.%Y"),line[4].strftime("%d.%m.%Y")]
        if oper == pfio:
            break
    if oper != pfio:
        print("\n!!! Пользователь {} не найден !!!\n")
        input("Нажмите любую клавишу чтобы закончить")
        exit()
    else:
        return pfio, name_work, dog_n, date_beg, date_end

# Функция поиска пользователя
def searchact(oper):
    indx = 0
    num, dakt, enakt, price = [],[],[],[]
    for row in results:
        line = row
        pfio, akt_number, dbeg_akt, dend_akt, price_pol = \
              [line[0],line[1],line[2],line[3],line[4]]
        if oper == pfio:
            num.append(akt_number)
            dakt.appent(dbeg_akt)
            endakt.append(dend_akt)
            price.append(price_pol)
            indx += 1
        else:
            continue
    date1 = datetime.datetime.strptime(endakt[indx-1].strftime("%d.%m.%Y"), "%d.%m.%Y")
    date1 = date(date1.year,date1.month+1,1)
    date2 = datetime.datetime.now()
    date2 = date(date2.year,date2.month,monthrange(date2.year,date2.month)[1])
    num_act = indx+1
    return num_act, date1, date2, price[indx-1]

# Функция преобрахования цифр в слова
def num2word(summ):
    ones = {"1":"один", "2":"два", "3":"три", "4":"четыре", "5":"пять", "6":"шесть", "7":"семь", "8":"восемь","9":"девять"}
    afterones = {"10":"десять","11":"одиннадцать", "12":"двенадцать", "13":"тринадцать", "14":"четырнадцать", \
                 "15":"пятнадцать", "16":"шестнадцать", "17":"семнадцать", "18":"восемнадцать","9":"девятнадцать"}
    tens = {"2":"двадцать", "3":"тридцать", "4":"сорок", "5":"пятьдесят", "6":"шестьдесят", "7":"семьдесят", \
            "8":"восемьдесят","9":"девяносто"}
    hundreds = {"1":"сто", "2":"двести", "3":"триста", "4":"четыреста", "5":"пятьсот", "6":"шестьсот", "7":"семьсот", \
                "8":"восемьсот","9":"девятьсот"}
    grand = {0:" миллиардов ", 1:" миллионов ", 2:" тысяч ", 3:""}

    def num_to_wrds(val):
        if val != "000":
            ans = ""
            if val[0] in hundreds:
                ans = ans + hundreds[val[0]] + " "
            if val[1:] in afterones:
                ans = ans + afterones[val[1:]] + " "
            elif val[1] in tens:
                if val[2] == "0":
                    ans = ans + tens[val[1]] + " "
                else: ans = ans + tens[val[1]] + " " + ones[val[2]]
            elif val[2] in ones:
                ans = ans + ones[val[2]]
            return ans

    num = summ
    pad = 12-len(str(num))
    numinwrds = "0"*pad + str(num)
    final = ""
    numlis = [numinwrds[0:3],numinwrds[3:6],numinwrds[6:9],numinwrds[9:12]]
    
    if len(str(num)) >= 4:
        for key,grp in enumerate(numlis):
            if grp != "000":
                final = final + num_to_wrds(grp) + grand[key]
                if final.startswith("один тысяч"):
                    final = final.replace("один тысяч", "одна тысяча")
                if final.startswith("два тысяч"):
                    final = final.replace("два тысяч", "две тысячи")
                if final.startswith("три тысяч") or final.startswith("четаре тысяч"):
                    final = final.replace("тысяч", "тысячи")
    else:
        final = final + num_to_wrds(num)
    return final

# НАЧАЛО РАБОТЫ

checkdate = False
mes = ["января","февраля","марта","апреля","мая","июня","июля","августа","сентбря","октября","ноября","декабря"]
while checkdate == False:
    dt = input('Введите дату в формате "ДД.ММ.ГГГГ":  ')
    if datetry(dt) == False:
        print ("Неправильно введена дата, попробуйте еще раз")
    else:
        checkdate = True
        day,month,year = map(int, dt.split("."))
        month = mes[month-1]
        dt = str(day)+" "+month+" "+str(year)+" г."


# Открываем таблицу "Оплата агентам за месяц.xlx"
wb = load_workbook(r"Оплата агентам за месяц.xlx")
sheet = wb["Лист 2"]
oper, passwd, mois,ep,vs,cost = ([],[],[],[],[],[])
list_param = [oper, passwd, mois,ep,vs,cost]
n = 0
df = pd.read_excel(r"Оплата агентам за месяц.xlx","Лист 2")
df = len(df)+1
for c in range (2,8):
    list1 = []
    for i in range (4,df):
        list1.append(sheet.cell(row = i, column = c).value)
        list_param[n].append(list1[i-4])
    n += 1

# Подключаемся к базе и получаем данные таблиц
file = open ("config")
var_dict = {}
for line in file.read().splitlines():
    var.append(line)
driver = '{}'.format(var[0])
server = '{}'.format(var[1])
port = '{}'.format(var[2])
db ='{}'.format(var[3])
user = '{}'.format(var[4])
pw = '{}'.format(var[5])
conn_str = ';'.join([driver, server, port, db, user, pw])
connetion = pyodbc.connect(conn_str)
coursor = connetion.cousor()
result = coursor.execute("SELECT pfio, akt_number, dbeg_akt, dend_akt, pricepol FROM OPERS_AKTS WHERE dog_date LIKE %2019% ORDER by akt_number").fetchall()
user = coursor.execute("SELECT pfio, name_work, dog_n, date_beg,date_end FROM OPERS WHERE date_beg LIKE %2019% ORDER by pfio").fetchall()

# Заполняем акты
df = df -2
for i in range (df):
    template = r"Акт сдачи-приеки.docx"
    document = MailMerge(template)
    signal = 0
    if "ИП " in oper[i]:
        signal = 1
        oper[i] = per[i][3:]
    pfio, name_work, dog_n, date_beg,date_end = search(oper[i])
    n_act, startD, endD, cost = searchact(oper[i])
    if signal == 1:
        name_work = "ИП " +name_work
        signal = 0
    date_act = dt
    operator = name_work
    print(operator)
    period = startD.strftime("%d.%m.%Y") +' - '+ endD.strftime("%d.%m.%Y")
    pwd = passwd[i]
    n_vs = n_ep = ""
    cvs = int(vs[i])
    cost_vs = float(cvs*cost)
    cep = int(ep[i])
    cost_ep = float(cep*cost)
    summ = int(cost_vs+cost_ep)
    sumword = num2word(str(summ))
    document.merge(
        NomEP = n_ep,
        DateDog = date_beg,
        Summ = str(toFix(summ,0)),
        NomVS = n_vs,
        NDog = dog_n,
        CostEP = str(toFix(cost_ep,2)),
        OperFull = operator,
        Passwd = pwd,
        NumAct = str(n_akt)
        CostVS = str(toFix(cost_vs,2)),
        DateAct = date_act,
        SummWord = sumword,
        DiaDate = period,
        OperSht = pfio)
    print(n_act, dog_n, date_beg, date_act, operator, period, pwd, n_vs, toFix(cost_vs,2), n_ep, toFix(cost_ep,2), toFix(summ,0), pfio, toFix(cost,2))
    document.write(r"Акт сдачи-приемки по оператору {}.docx".format(oper[i]))
